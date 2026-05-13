"""Medusa CMMS Excel adapter for importing work-order data.

This module is the CMMS integration seam.  The current implementation reads
an Excel export produced by the Medusa CMMS system used at Helse Stavanger
and normalises it into a list of dicts with English field names.  Norwegian
column headers and work-order type labels are translated via module-level
mapping tables (``_COLUMN_MAP``, ``_TYPE_MAP``).

The adapter is designed for easy replacement: once a REST API is available
the public interface (``parse_medusa_excel``) can be replaced by an API
client without any changes to calling code.
"""

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from ventilator_pdm.registry import REG_TO_SERIAL

# Norwegian → internal column mapping
_COLUMN_MAP = {
    "AO-nr.": "work_order_id",
    "AO-type": "work_order_type",
    "Reg. dato": "registered_date",
    "Ferdig": "completed_date",
    "Oppgave/feilbeskrivelse": "fault_description",
    "Teknisk beskrivelse": "technical_description",
    "Reg.nr.": "device_reg",
    "Serienr.": "serial_number",
}

# AO-type Norwegian → English mapping
_TYPE_MAP = {
    "korrektiv": "corrective",
    "forebyggende": "preventive",
}


def _normalize_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        return date.fromisoformat(val[:10])
    return None


def _normalize_type(ao_type: str) -> str:
    return _TYPE_MAP.get(ao_type.strip().lower(), ao_type.strip().lower())



def parse_medusa_excel(path: Path) -> list[dict]:
    """Parse a Medusa CMMS Excel export into normalised work-order records.

    Reads the first (active) sheet of the workbook.  The first row is
    expected to contain Norwegian column headers that are mapped to English
    field names via ``_COLUMN_MAP``.  Rows without a device registration
    number are skipped.  Work-order type values are translated from
    Norwegian to English via ``_TYPE_MAP``; unknown types are passed
    through in lower-case.  Date fields are normalised to
    :class:`datetime.date` objects via ``_normalize_date``.  Device serial
    numbers are resolved from the registration number using
    :data:`pdm.registry.REG_TO_SERIAL`; unresolved devices have
    ``device_serial`` set to ``None``.

    Args:
        path: Filesystem path to the ``.xlsx`` file exported from Medusa.

    Returns:
        List of work-order dicts, one per data row.  Each dict contains:

        - ``work_order_id`` (str)
        - ``work_order_type`` (str): ``"corrective"``, ``"preventive"``,
            or the raw lower-cased AO-type value.
        - ``registered_date`` (:class:`datetime.date` | ``None``)
        - ``completed_date`` (:class:`datetime.date` | ``None``)
        - ``fault_description`` (str)
        - ``technical_description`` (str)
        - ``device_reg`` (int): CMMS registration number.
        - ``serial_number`` (str): Raw serial field from CMMS (may be
            empty if the column is absent).
        - ``device_serial`` (str | ``None``): Fleet serial resolved via
            registry, or ``None`` if not found.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header = [str(c).strip() if c else "" for c in rows[0]]
    col_indices = {}
    for no_col, name in _COLUMN_MAP.items():
        if no_col in header:
            col_indices[name] = header.index(no_col)

    records = []
    for row in rows[1:]:
        device_reg = row[col_indices["device_reg"]]
        if device_reg is None:
            continue  # Skip blank rows
        if isinstance(device_reg, str):
            device_reg = int(device_reg)

        serial_nr = row[col_indices["serial_number"]] if "serial_number" in col_indices else None

        records.append({
            "work_order_id": str(row[col_indices["work_order_id"]]),
            "work_order_type": _normalize_type(str(row[col_indices["work_order_type"]] or "")),
            "registered_date": _normalize_date(row[col_indices["registered_date"]]),
            "completed_date": _normalize_date(row[col_indices["completed_date"]]),
            "fault_description": str(row[col_indices["fault_description"]] or ""),
            "technical_description": str(row[col_indices["technical_description"]] or ""),
            "device_reg": device_reg,
            "serial_number": str(serial_nr) if serial_nr else "",
            "device_serial": REG_TO_SERIAL.get(device_reg),
        })

    return records
